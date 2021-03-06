#!/usr/bin/python
# -*- coding: UTF-8 -*- 

from openpyxl.styles import Font, Border, Side, Alignment
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr
import pymysql
import logging
from didi_order_check import config


'''拉取订单数据'''
def pull_data(sub_status):
    try:
        s = []
        conn = pymysql.connect(**config.DB_CONFIG_BUY)
        cursor = conn.cursor()
        cursor.execute(config.SQL_order_price%(sub_status))
        results = cursor.fetchall()
        for x in list(results):
            x = list(x)
            if x[1] == 0:
                x[1] = '因公订单'
            elif x[1] == 8:
                x[1] = '积分订单'
            elif x[1] == 13:
                x[1] = '彩豆订单'
            else:
                x[1] = '无效订单'
            s.append(x)
    except:
        logging.error("DB连接错误", exc_info=True)

    finally:
        conn.close()
        return s

'''连接DB'''
def db_connect(db_config):
    try:
        conn = pymysql.connect(**db_config)
        return conn
    except:
        logging.error("DB连接错误", exc_info=True)


'''断开链接DB'''
def db_disconnect(conn):
    try:
        conn.close()
    except:
        logging.error("DB连接错误", exc_info=True)

'''拉取db库数据'''
def pull_data(conn, table_list, parameter):
    try:
        s = []
        cursor = conn.cursor(cursor=pymysql.cursors.DictCursor)
        for x in table_list:
            cursor.execute(x%(parameter, ))
            # data_title = []
            # for field in cursor.description:
            #     data_title.append(field[0])
            result = cursor.fetchall()
            # for x in result:
            #     data_dict = {}
            #     for y in range(len(x)):
            #         data_dict[data_title[y]] = x[y]
            #     s.append(data_dict)
            s.append(result)
        # print(s)
        return s

    except:
        cursor.close()
        logging.error("DB连接错误", exc_info=True)


'''Excel表格格式优化'''
def sheet_layout(ws):
    # 设置单元格字体为11号黑体
    font = Font(name='黑体', size=11)
    # 设置单元格格式为水平居中，垂直居中
    alignment = Alignment(horizontal='center', vertical='center')
    # 设置单元格边框颜色为黑色
    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    for column in ws.columns:
        for cell in column:
            cell.alignment = alignment
            cell.font = font
            cell.border = border

'''提取Excel数据到字典'''
# def pull(ws):
#     dict = {}
#     for x in range(2, ws.max_row + 1):
#         dict[ws.cell(x, 1).value] = ws.cell(x, 2).value
#     return dict

'''发送邮件'''
def mail(my_user, contant, title):
    my_sender = 'zhangt@shinemo.com'  # 发件人邮箱账号，为了后面易于维护，所以写成了变量
    try:
        msg = MIMEText(contant, 'plain', 'utf-8')
        msg['From'] = formataddr(["系统通知", my_sender])  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
        msg['Subject'] = title  # 邮件的主题，也可以说是标题
        server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)  # 发件人邮箱中的SMTP服务器，端口是25
        server.login(my_sender, "12345qWE")  # 括号中对应的是发件人邮箱账号、邮箱密码
        server.sendmail(my_sender, my_user, msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
        server.quit()  # 这句是关闭连接的意思

    except:
        print('邮件发送失败！')